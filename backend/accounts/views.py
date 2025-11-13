import openpyxl
from django.http import HttpResponse
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.utils import timezone
from datetime import datetime, timedelta, time
from django.conf import settings
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.timezone import now, make_aware, timezone
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import (
    Count,
    Avg,
    F,
    DateField,
    Q,
    Sum,
    F,
    ExpressionWrapper,
    DurationField,
)
from django.db.models.functions import TruncDay
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from decouple import config
from rest_framework import status, generics, permissions, viewsets, filters
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
import os
from .permissions import IsAdministrativo


# Models
from .models import (
    Producto,
    OrdenItem,
    Notificacion,
    Orden,
    Agendamiento,
    Vehiculo,
    OrdenHistorialEstado,
    OrdenPausa,
    OrdenDocumento,
    LlaveVehiculo,
    PrestamoLlave,
    LlaveHistorialEstado,
    Taller,
    Usuario,
    AgendamientoHistorial
)

# Serializers
from .serializers import (
    ProductoSerializer,
    OrdenItemSerializer,
    NotificacionSerializer,
    OrdenSerializer,
    AgendamientoSerializer,
    OrdenDocumentoSerializer,
    VehiculoSerializer,
    LoginSerializer,
    UserSerializer,
    ChangePasswordSerializer,
    UserCreateUpdateSerializer,
    LlaveVehiculoSerializer,
    PrestamoLlaveSerializer,
    LlaveHistorialEstadoSerializer,
    HistorialSeguridadSerializer,
    OrdenSalidaListSerializer,
    TallerSerializer,
)

User = get_user_model()


def enviar_correo_notificacion(usuario, subject, message_body):
    """
    Envía un correo electrónico de notificación usando la plantilla HTML.
    """
    # 1. Asegurarnos que el usuario tenga un email
    if not usuario.email:
        print(f"Usuario {usuario.username} no tiene email, no se envía correo.")
        return

    # 2. ***** CONFIGURACIÓN DE PRUEBA *****
    recipient_email = "fer.araneda@duocuc.cl"
    print(
        f"Enviando correo de prueba a: {recipient_email} (Usuario real: {usuario.email})"
    )

    # 3. Preparar el contexto para la plantilla HTML
    context = {
        "subject": subject,
        "message_body": message_body,
        "nombre_usuario": usuario.first_name or usuario.username,
    }

    try:

        html_message = render_to_string("emails/notificacion_base.html", context)

        plain_message = strip_tags(html_message)

        send_mail(
            subject,
            plain_message,
            None,
            [recipient_email],
            html_message=html_message,
            fail_silently=False,
        )
        print(f"Correo enviado exitosamente a {recipient_email}")

    except Exception as e:
        print(f"ERROR al enviar correo a {recipient_email}: {e}")


# --------------------
# Permisos personalizados
# --------------------
class IsSupervisor(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(
            request.user
            and request.user.is_authenticated
            and request.user.groups.filter(
                name__in=["Supervisor", "Administrativo"]
            ).exists()
        )


class IsSupervisorOrMecanico(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(
            request.user
            and request.user.is_authenticated
            and request.user.groups.filter(
                name__in=["Supervisor", "Mecanico", "Administrativo"]
            ).exists()
        )


class IsSupervisorOrSeguridad(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(
            request.user
            and request.user.is_authenticated
            and request.user.groups.filter(
                name__in=["Supervisor", "Seguridad", "Administrativo"]
            ).exists()
        )


class IsAdministrativo(permissions.BasePermission):
    """
    Permiso específico para el rol Administrativo o Supervisor.
    """

    def has_permission(self, request, view):
        return bool(
            request.user
            and request.user.is_authenticated
            and request.user.groups.filter(
                name__in=["Administrativo", "Supervisor"]
            ).exists()
        )


class IsControlLlaves(permissions.BasePermission):
    """
    Permiso para el Encargado de Llaves o Supervisor.
    """

    def has_permission(self, request, view):
        return bool(
            request.user
            and request.user.is_authenticated
            and request.user.groups.filter(
                name__in=["Control Llaves", "Supervisor", "Administrativo"]
            ).exists()
        )


class IsSupervisorOrControlLlaves(permissions.BasePermission):
    """
    Permiso para Supervisor O Encargado de Llaves (para ver listas de usuarios).
    """

    def has_permission(self, request, view):
        return bool(
            request.user
            and request.user.is_authenticated
            and request.user.groups.filter(
                name__in=["Supervisor", "Control Llaves", "Administrativo"]
            ).exists()
        )


class IsRepuestos(permissions.BasePermission):
    """
    Permiso para el rol de Repuestos o Supervisor.
    """

    def has_permission(self, request, view):
        return bool(
            request.user
            and request.user.is_authenticated
            and request.user.groups.filter(
                name__in=["Repuestos", "Supervisor", "Administrativo"]
            ).exists()
        )


# --------------------
# Autenticación y perfil
# --------------------
class LoginView(generics.GenericAPIView):
    serializer_class = LoginSerializer
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data["user"]
        refresh = RefreshToken.for_user(user)
        user_data = UserSerializer(user).data
        return Response(
            {
                "refresh": str(refresh),
                "access": str(refresh.access_token),
                "user": user_data,
            },
            status=status.HTTP_200_OK,
        )


class PasswordResetRequestView(generics.GenericAPIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get("email")
        if not email:
            return Response(
                {"error": "Se requiere el correo"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            # No revelar si el correo existe o no
            return Response(
                {
                    "message": "Si el correo está registrado, se enviará un enlace de recuperación."
                },
                status=status.HTTP_200_OK,
            )

        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        frontend = config("FRONTEND_URL", default="http://localhost:5173")
        reset_link = f"{frontend.rstrip('/')}/set-new-password?uid={uid}&token={token}"

        send_mail(
            "Restablecer contraseña para Taller PepsiCo",
            f"Hola {user.first_name},\n\nUsa este enlace para restablecer tu contraseña: {reset_link}\n\nSi no solicitaste esto, ignora este mensaje.",
            None,
            [email],
            fail_silently=False,
        )
        return Response(
            {
                "message": "Si el correo está registrado, se enviará un enlace de recuperación."
            },
            status=status.HTTP_200_OK,
        )


class PasswordResetConfirmView(generics.GenericAPIView):
    permission_classes = [AllowAny]

    def post(self, request):
        uidb64 = request.data.get("uid")
        token = request.data.get("token")
        new_password = request.data.get("password")

        if not uidb64 or not token or not new_password:
            return Response(
                {"error": "Datos incompletos"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            uid_decoded = urlsafe_base64_decode(uidb64).decode()
            uid = int(uid_decoded)
            user = User.objects.get(pk=uid)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            return Response(
                {"error": "El enlace de restablecimiento es inválido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not default_token_generator.check_token(user, token):
            return Response(
                {"error": "El enlace de restablecimiento es inválido o ha expirado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            validate_password(new_password, user)
        except ValidationError as e:
            return Response(
                {"error": list(e.messages)}, status=status.HTTP_400_BAD_REQUEST
            )

        user.set_password(new_password)
        user.save()
        return Response(
            {"message": "Contraseña restablecida con éxito"}, status=status.HTTP_200_OK
        )


class UserDetailView(generics.RetrieveAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user


class ChangePasswordView(generics.GenericAPIView):
    serializer_class = ChangePasswordSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = request.user
        old_password = serializer.validated_data["old_password"]
        new_password = serializer.validated_data["new_password"]

        if not user.check_password(old_password):
            return Response(
                {"error": "La contraseña actual es incorrecta."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            validate_password(new_password, user)
        except ValidationError as e:
            return Response(
                {"error": list(e.messages)}, status=status.HTTP_400_BAD_REQUEST
            )

        user.set_password(new_password)
        user.save()
        return Response(
            {"message": "Contraseña cambiada con éxito."}, status=status.HTTP_200_OK
        )


# --------------------
# Gestión de usuarios (Supervisor)
# --------------------
class UserListView(generics.ListAPIView):
    queryset = User.objects.all().order_by("first_name")
    serializer_class = UserSerializer
    permission_classes = [IsSupervisorOrControlLlaves]


class UserCreateAPIView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserCreateUpdateSerializer
    permission_classes = [IsSupervisor]


class UserRetrieveUpdateAPIView(generics.RetrieveUpdateAPIView):
    queryset = User.objects.all()
    serializer_class = UserCreateUpdateSerializer
    permission_classes = [IsSupervisor]
    lookup_field = "id"

    def get_serializer_class(self):
        """
        Usa un serializador diferente para LEER (GET) que para ESCRIBIR (PUT/PATCH).
        """
        if self.request.method in ["PUT", "PATCH"]:
            # Al actualizar, usamos el serializador que acepta el 'rol' por texto.
            return UserCreateUpdateSerializer

        # Al cargar (GET), usamos el serializador que SÍ muestra el 'rol'.
        return UserSerializer


class ChoferListView(generics.ListAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return User.activos.filter(groups__name="Chofer").order_by("first_name")


# --------------------
# Dashboard supervisor
# --------------------
dias_semana = {0: "Lun", 1: "Mar", 2: "Mié", 3: "Jue", 4: "Vie", 5: "Sáb", 6: "Dom"}


@api_view(["GET"])
@permission_classes([IsSupervisor])
def supervisor_dashboard_stats(request):
    today = now().date()
    start_of_month_dt = timezone.make_aware(
        datetime.combine(today.replace(day=1), datetime.min.time())
    )
    start_of_week = today - timedelta(days=today.weekday())
    start_of_week_dt = timezone.make_aware(
        datetime.combine(start_of_week, datetime.min.time())
    )

    pendientes_aprobacion = Agendamiento.objects.filter(
        estado=Agendamiento.Estado.PROGRAMADO
    ).count()

    # Vehículos en taller (fallback si no existe manager 'activas')
    try:
        vehiculos_en_taller = (
            Orden.objects.activas().values("vehiculo").distinct().count()
        )
    except Exception:
        # Fallback si 'activas' no existe o falla
        vehiculos_en_taller = (
            Orden.objects.exclude(estado=Orden.Estado.FINALIZADO)
            .values("vehiculo")
            .distinct()
            .count()
        )

    # Agendamientos para HOY (usamos CONFIRMADO para alinearnos con SeguridadAgendaView)
    start_today = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    end_today = start_today + timedelta(days=1)
    agendamientos_hoy = Agendamiento.objects.filter(
        estado=Agendamiento.Estado.CONFIRMADO,  # Correcto
        fecha_hora_programada__gte=start_today,
        fecha_hora_programada__lt=end_today,
    ).count()

    # Órdenes finalizadas este mes
    ordenes_finalizadas_mes = Orden.objects.filter(
        estado=Orden.Estado.FINALIZADO,  # Correcto
        fecha_entrega_real__gte=start_of_month_dt,
    ).count()

    # Tiempo promedio de reparación (en días, con manejo de nulls)
    ordenes_completadas = Orden.objects.filter(
        estado=Orden.Estado.FINALIZADO,
        fecha_entrega_real__isnull=False,
        fecha_ingreso__isnull=False,
    )
    tiempo_promedio_str = "N/A"
    if ordenes_completadas.exists():
        avg_delta = ordenes_completadas.aggregate(
            avg_duration=Avg(F("fecha_entrega_real") - F("fecha_ingreso"))
        )["avg_duration"]
        if avg_delta:
            total_dias = avg_delta.total_seconds() / (60 * 60 * 24)
            tiempo_promedio_str = f"{total_dias:.1f} días"

    # Órdenes por estado
    ordenes_por_estado = list(
        Orden.objects.values("estado").annotate(cantidad=Count("id")).order_by("estado")
    )

    # Órdenes última semana (por día)
    ordenes_semana_raw = (
        Orden.objects.filter(fecha_ingreso__gte=start_of_week_dt)
        .annotate(dia_semana=TruncDay("fecha_ingreso", output_field=DateField()))
        .values("dia_semana")
        .annotate(creadas=Count("id"))
        .order_by("dia_semana")
    )
    ordenes_ultima_semana = []
    dias_semana_map = {
        0: "Lun",
        1: "Mar",
        2: "Mié",
        3: "Jue",
        4: "Vie",
        5: "Sáb",
        6: "Dom",
    }
    for i in range(7):
        fecha_dia = start_of_week + timedelta(days=i)
        dia_nombre = dias_semana_map.get(fecha_dia.weekday(), "")
        cre = 0
        for item in ordenes_semana_raw:
            if item["dia_semana"] == fecha_dia:
                cre = item["creadas"]
                break
        ordenes_ultima_semana.append({"dia": dia_nombre, "creadas": cre})

    ordenes_recientes = list(
        Orden.objects.select_related("vehiculo", "usuario_asignado")
        .order_by("-fecha_ingreso")[:10]
        .values(
            "id",
            "vehiculo__patente",
            "estado",
            "usuario_asignado__first_name",
            "usuario_asignado__last_name",
            "usuario_asignado__username",
        )
    )
    ordenes_recientes_data = []
    for o in ordenes_recientes:
        first_name = o.get("usuario_asignado__first_name") or ""
        last_name = o.get("usuario_asignado__last_name") or ""
        username = o.get("usuario_asignado__username") or ""
        mecanico_nombre = (
            f"{first_name} {last_name}".strip() or username or "No asignado"
        )
        ordenes_recientes_data.append(
            {
                "id": o["id"],
                "patente": o.get("vehiculo__patente") or "Sin patente",
                "estado": o["estado"],
                "mecanico": mecanico_nombre,
            }
        )

    response_data = {
        "kpis": {
            "vehiculosEnTaller": vehiculos_en_taller,
            "agendamientosHoy": agendamientos_hoy,
            "ordenesFinalizadasMes": ordenes_finalizadas_mes,
            "tiempoPromedioRep": tiempo_promedio_str,
        },
        "alertas": {
            "pendientesAprobacion": pendientes_aprobacion,
        },
        "ordenesPorEstado": ordenes_por_estado,
        "ordenesUltimaSemana": ordenes_ultima_semana,
        "ordenesRecientes": ordenes_recientes_data,
    }
    return Response(response_data, status=status.HTTP_200_OK)


# --------------------
# ViewSets
# --------------------
class VehiculoViewSet(viewsets.ModelViewSet):
    serializer_class = VehiculoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.action in [
            "retrieve",
            "update",
            "partial_update",
            "destroy",
            "reactivar",
        ]:
            return Vehiculo.objects.all()
        user = self.request.user
        if user.groups.filter(name="Chofer").exists():
            return Vehiculo.activos.filter(chofer=user)
        return Vehiculo.activos.all()

    def perform_create(self, serializer):
        """
        Se ejecuta después de crear un Vehículo.
        Automatiza la creación de sus llaves por defecto.
        """

        vehiculo = serializer.save()

        try:

            LlaveVehiculo.objects.get_or_create(
                vehiculo=vehiculo,
                tipo=LlaveVehiculo.Tipo.ORIGINAL,
                defaults={"codigo_interno": f"{vehiculo.patente}-ORI"},
            )

            LlaveVehiculo.objects.get_or_create(
                vehiculo=vehiculo,
                tipo=LlaveVehiculo.Tipo.DUPLICADO,
                defaults={"codigo_interno": f"{vehiculo.patente}-DUP"},
            )
        except Exception as e:

            print(
                f"ERROR: No se pudieron crear llaves automáticas para {vehiculo.patente}: {e}"
            )

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

    @action(
        detail=False,
        methods=["get"],
        url_path="inactivos",
        permission_classes=[IsAuthenticated],
    )
    def inactivos(self, request):
        vehiculos_inactivos = Vehiculo.objects.filter(is_active=False)
        user = self.request.user
        if user.groups.filter(name="Chofer").exists():
            vehiculos_inactivos = vehiculos_inactivos.filter(chofer=user)
        serializer = self.get_serializer(vehiculos_inactivos, many=True)
        return Response(serializer.data)

    @action(
        detail=True,
        methods=["post"],
        url_path="reactivar",
        permission_classes=[IsSupervisor],
    )
    def reactivar(self, request, pk=None):
        vehiculo = self.get_object()
        vehiculo.is_active = True
        vehiculo.save()
        return Response(self.get_serializer(vehiculo).data, status=status.HTTP_200_OK)


class AgendamientoViewSet(viewsets.ModelViewSet):
    serializer_class = AgendamientoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(
            name__in=["Supervisor", "Mecanico", "Seguridad", "Administrativo"]
        ).exists():
            return (
                Agendamiento.objects.select_related("vehiculo", "mecanico_asignado")
                .all()
                .order_by("fecha_hora_programada")
            )
        elif user.groups.filter(name="Chofer").exists():
            return Agendamiento.objects.filter(vehiculo__chofer=user).order_by(
                "fecha_hora_programada"
            )
        return Agendamiento.objects.none()

    def perform_create(self, serializer):
        user = self.request.user

        # 1. Guardamos la instancia UNA SOLA VEZ y la asignamos a una variable
        agendamiento = serializer.save(creado_por=user, chofer_asociado=user)

        # 2. Ahora usamos esa variable 'agendamiento' para las notificaciones
        try:
            # 2. Obtenemos a todos los supervisores activos
            supervisores = User.objects.filter(
                groups__name__in=["Supervisor", "Administrativo"], is_active=True
            )

            # 3. Preparamos el mensaje
            chofer = agendamiento.creado_por
            chofer_nombre = (
                f"{chofer.first_name} {chofer.last_name}".strip() or chofer.username
            )
            patente = agendamiento.vehiculo.patente

            subject = f"Nueva Solicitud de Cita: {patente}"
            mensaje = f"El chofer {chofer_nombre} ha solicitado un ingreso para el vehículo {patente}. Motivo: {agendamiento.motivo_ingreso}"
            link_supervisor = "/panel-supervisor"  # Link al panel donde aprueban

            # 4. Enviamos notificación y email a CADA supervisor
            for supervisor in supervisores:
                Notificacion.objects.create(
                    usuario=supervisor, mensaje=mensaje, link=link_supervisor
                )

                # 5. Enviamos el correo
                enviar_correo_notificacion(supervisor, subject, mensaje)

        except Exception as e:

            print(f"ERROR al notificar al supervisor sobre nueva cita: {e}")

    @action(
        detail=True,
        methods=["get"],
        url_path="verificar-stock-mantenimiento",
        permission_classes=[IsSupervisor],  # Solo el supervisor puede ver esto
    )
    def verificar_stock_mantenimiento(self, request, pk=None):
        agendamiento = self.get_object()

        if not agendamiento.es_mantenimiento:
            return Response(
                {"error": "Esta no es una cita de mantenimiento."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # --- LISTA DE REPUESTOS PREDETERMINADOS ---
        # ¡IMPORTANTE! Debes usar los SKUs reales de tu BD.
        # Usaré los de tu archivo populate_data.py como ejemplo.
        REPUESTOS_MANTENIMIENTO = {
            "ACE-10W40": 1,  # SKU: Cantidad necesaria
            "FIL-AIRE-01": 1,
            "FRE-LIQ-01": 1,
            # 'SKU-FILTRO-ACEITE': 1, # <--- Añade más aquí
        }

        skus_requeridos = list(REPUESTOS_MANTENIMIENTO.keys())
        productos = Producto.objects.filter(sku__in=skus_requeridos)

        faltantes = []
        stock_completo = True

        for sku, cantidad_necesaria in REPUESTOS_MANTENIMIENTO.items():
            try:
                producto = productos.get(sku=sku)
                if producto.stock < cantidad_necesaria:
                    stock_completo = False
                    faltantes.append(
                        {
                            "nombre": producto.nombre,
                            "sku": sku,
                            "stock_actual": producto.stock,
                            "necesario": cantidad_necesaria,
                        }
                    )
            except Producto.DoesNotExist:
                stock_completo = False
                faltantes.append(
                    {
                        "nombre": f"Producto (SKU: {sku}) no encontrado",
                        "sku": sku,
                        "stock_actual": 0,
                        "necesario": cantidad_necesaria,
                    }
                )

        return Response(
            {
                "stock_completo": stock_completo,
                "faltantes": faltantes,
                "repuestos_revisados": REPUESTOS_MANTENIMIENTO,
            },
            status=status.HTTP_200_OK,
        )

    @action(
        detail=True,
        methods=["post"],
        url_path="confirmar-y-asignar",
        permission_classes=[IsSupervisor],
    )
    def confirmar_y_asignar(self, request, pk=None):
        with transaction.atomic():
            agendamiento = self.get_object()
            mecanico_id_raw = request.data.get("mecanico_id")
            fecha_hora_asignada_str = request.data.get("fecha_hora_asignada")
            motivo_cambio = request.data.get("motivo_reagendamiento", None)

            REPUESTOS_MANTENIMIENTO = {}
            if agendamiento.es_mantenimiento:
                REPUESTOS_MANTENIMIENTO = {
                    "ACE-10W40": 1,
                    "FIL-AIRE-01": 1,
                    "FRE-LIQ-01": 1,  # <-- SKU CORREGIDO
                }
            try:
                mecanico_id = int(mecanico_id_raw)
                mecanico = User.objects.get(id=mecanico_id, groups__name="Mecanico")
            except (TypeError, ValueError, User.DoesNotExist):
                return Response(
                    {"error": "El mecánico seleccionado es inválido."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            fecha_a_validar = None
            hubo_cambio_fecha = False

            if not fecha_hora_asignada_str:

                return Response(
                    {"error": "Debe seleccionar una fecha y hora."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:
                # fromisoformat() ya maneja el string 'Z' (UTC) de JavaScript
                fecha_a_validar = datetime.fromisoformat(fecha_hora_asignada_str)

                # Comparamos la nueva fecha con la original (si existía)
                if agendamiento.fecha_hora_programada != fecha_a_validar:
                    hubo_cambio_fecha = True

            except (ValueError, TypeError):
                return Response(
                    {"error": "Formato de fecha/hora asignada es inválido."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # --- INICIO BLOQUE DE VALIDACIÓN DE STOCK MEJORADO ---
            if agendamiento.es_mantenimiento:
                skus_requeridos = list(REPUESTOS_MANTENIMIENTO.keys())
                # Bloqueamos los productos para la transacción
                productos = Producto.objects.select_for_update().filter(
                    sku__in=skus_requeridos
                )

                # Convertimos la consulta a un diccionario para mejor acceso
                productos_encontrados = {p.sku: p for p in productos}

                for sku, cantidad_necesaria in REPUESTOS_MANTENIMIENTO.items():

                    producto = productos_encontrados.get(sku)

                    # 1. Verificación: ¿Se encontró el producto?
                    if not producto:
                        # Este es el error 'DoesNotExist'.
                        return Response(
                            {
                                "error": f"Producto con SKU '{sku}' no fue encontrado en la base de datos. Cita no confirmada."
                            },
                            status=status.HTTP_404_NOT_FOUND,
                        )

                    # 2. Verificación: ¿Hay stock?
                    if producto.stock < cantidad_necesaria:
                        # ¡Este es el error que SÍ muestra el nombre!
                        return Response(
                            {
                                "error": f"Stock agotado para '{producto.nombre}' (SKU: {sku}). Quedan {producto.stock}. Cita no confirmada."
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )

        # Calculamos el fin
        fecha_fin = fecha_a_validar + timedelta(
            minutes=agendamiento.duracion_estimada_minutos
        )

        # 4. VALIDACIÓN DE CONFLICTO (MECÁNICO)
        overlapping_mecanico = Agendamiento.objects.filter(
            Q(fecha_hora_programada__lt=fecha_fin)
            & Q(fecha_hora_fin__gt=fecha_a_validar)
            & Q(mecanico_asignado=mecanico)
            & Q(
                estado__in=[
                    Agendamiento.Estado.CONFIRMADO,
                    Agendamiento.Estado.EN_TALLER,
                ]
            )
        ).exclude(pk=agendamiento.pk)

        if overlapping_mecanico.exists():
            return Response(
                {
                    "error": f"Conflicto de horario (Mecánico): El mecánico {mecanico.get_full_name()} ya tiene una cita en ese rango."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 5. VALIDACIÓN DE CONFLICTO (VEHÍCULO)
        overlapping_vehiculo = Agendamiento.objects.filter(
            Q(vehiculo=agendamiento.vehiculo)
            & Q(fecha_hora_programada__lt=fecha_fin)
            & Q(fecha_hora_fin__gt=fecha_a_validar)
            & ~Q(estado__in=["Finalizado", "Cancelado"])  # <--- Lógica de la constraint
        ).exclude(pk=agendamiento.pk)

        if overlapping_vehiculo.exists():
            return Response(
                {
                    "error": f"Conflicto de horario (Vehículo): El vehículo {agendamiento.vehiculo.patente} ya tiene OTRA cita activa en ese nuevo rango."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 6. GUARDAR TODO (SIN chofer_asociado)
        agendamiento.mecanico_asignado = mecanico
        agendamiento.estado = Agendamiento.Estado.CONFIRMADO
        agendamiento.fecha_hora_programada = fecha_a_validar  # Asignamos la nueva fecha

        if hubo_cambio_fecha:
            agendamiento.motivo_reagendamiento = motivo_cambio

            # Notificar al chofer (Esta lógica está bien)
            try:
                if agendamiento.chofer_asociado:
                    # --- 1. Obtenemos la dirección del taller ---
                    taller_direccion = (
                        "Taller no especificado. Consulte con su supervisor."
                    )
                    if agendamiento.vehiculo and agendamiento.vehiculo.taller:
                        taller_direccion = agendamiento.vehiculo.taller.direccion

                    # --- 2. Creamos el mensaje con la dirección ---
                    fecha_str = fecha_a_validar.strftime("%d-%m-%Y a las %H:%M")
                    motivo_str = motivo_cambio or "Asignación de taller."
                    mensaje = (
                        f"Cita Confirmada: Su cita para {agendamiento.vehiculo.patente} es el {fecha_str}. "
                        f"Dirección Taller: {taller_direccion}. Motivo: {motivo_str}"
                    )

                    # --- 3. Enviamos la notificación y el correo ---
                    Notificacion.objects.create(
                        usuario=agendamiento.chofer_asociado,
                        mensaje=mensaje,
                        link=f"/historial",  # Link al historial del chofer
                    )
                    subject_chofer = f"Cita Confirmada: {agendamiento.vehiculo.patente} el {fecha_str}"
                    enviar_correo_notificacion(
                        agendamiento.chofer_asociado, subject_chofer, mensaje
                    )
            except Exception as e:
                print(f"Error al crear notificación de reagendamiento: {e}")

            try:
                # 1. Buscamos a todos los usuarios del grupo "Seguridad"
                usuarios_seguridad = User.objects.filter(
                    groups__name="Seguridad", is_active=True
                )

                # 2. Creamos el mensaje
                mensaje_seguridad = f"Vehículo {agendamiento.vehiculo.patente} (Chofer: {agendamiento.chofer_asociado.first_name}) tiene cita confirmada para el {fecha_a_validar.strftime('%d-%m a las %H:%M')}."

                # 3. Creamos una notificación para cada uno de ellos
                for user_seg in usuarios_seguridad:
                    Notificacion.objects.create(
                        usuario=user_seg,
                        mensaje=mensaje_seguridad,
                        link="/panel-ingresos",  # El link a su panel de trabajo
                    )
                    subject_seguridad = (
                        f"Cita Confirmada: Vehículo {agendamiento.vehiculo.patente}"
                    )
                    enviar_correo_notificacion(
                        user_seg, subject_seguridad, mensaje_seguridad
                    )
            except Exception as e:
                # Si falla la notificación de seguridad, no detenemos la operación
                print(f"Error al crear notificación para Seguridad: {e}")

            if agendamiento.es_mantenimiento:
                # 1. Crear la Orden de Trabajo asociada
                nueva_orden = Orden.objects.create(
                    vehiculo=agendamiento.vehiculo,
                    agendamiento_origen=agendamiento,
                    descripcion_falla=agendamiento.motivo_ingreso,
                    usuario_asignado=agendamiento.mecanico_asignado,
                    estado=Orden.Estado.INGRESADO,  # Queda lista para cuando el auto llegue
                )

                # 2. Descontar stock y crear los items (ya pre-aprobados)
                for sku, cantidad_necesaria in REPUESTOS_MANTENIMIENTO.items():
                    # Usamos 'productos.get()' del 'select_for_update()' que hicimos arriba
                    producto = productos.get(sku=sku)

                    # 2a. Descontar Stock
                    producto.stock = F("stock") - cantidad_necesaria
                    producto.save()

                    # 2b. Crear el OrdenItem como APROBADO
                    OrdenItem.objects.create(
                        orden=nueva_orden,
                        producto=producto,
                        cantidad=cantidad_necesaria,
                        precio_unitario=producto.precio_venta,
                        solicitado_por=request.user,  # Solicitado por el Supervisor
                        gestionado_por=request.user,  # Aprobado por el Supervisor
                        fecha_gestion=timezone.now(),  # Aprobado ahora mismo
                        estado_repuesto=OrdenItem.EstadoRepuesto.APROBADO,  # ¡Estado Aprobado!
                    )

                    AgendamientoHistorial.objects.create(
                        agendamiento=agendamiento,
                        estado=agendamiento.estado,  # Guardará "Confirmado"
                        usuario=request.user,
                        comentario=motivo_cambio or "Cita confirmada y asignada.",
                    )

        agendamiento.save()  # Esta línea ya no dará error
        return Response(
            self.get_serializer(agendamiento).data, status=status.HTTP_200_OK
        )

    @action(
        detail=True,
        methods=["post"],
        url_path="registrar-ingreso",
        permission_classes=[IsSupervisorOrSeguridad],
    )
    @transaction.atomic  # <--- Añadimos transacción por seguridad
    def registrar_ingreso(self, request, pk=None):
        agendamiento = self.get_object()
        if agendamiento.estado != Agendamiento.Estado.CONFIRMADO:
            return Response(
                {"error": "Solo se puede registrar el ingreso de una cita confirmada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # --- LÓGICA MODIFICADA ---
        # 1. Buscamos si la orden YA EXISTE (creada por el supervisor)
        try:
            # Usamos 'orden_generada' (el related_name del OneToOneField)
            nueva_orden = agendamiento.orden_generada

            # Si existe, solo actualizamos el estado (si es necesario)
            if nueva_orden.estado == Orden.Estado.INGRESADO:
                # (Opcional: puedes cambiarlo a 'En Diagnóstico' si quieres)
                pass  # Ya está 'Ingresado', no hacemos nada.

            mensaje_respuesta = "Ingreso registrado en orden existente."

        except Orden.DoesNotExist:
            # 2. Si NO existe (cita normal), la CREAMOS.
            nueva_orden = Orden.objects.create(
                vehiculo=agendamiento.vehiculo,
                agendamiento_origen=agendamiento,
                descripcion_falla=agendamiento.motivo_ingreso,
                usuario_asignado=agendamiento.mecanico_asignado,
                estado=Orden.Estado.INGRESADO,  # <--- Estado inicial
            )
            mensaje_respuesta = "Ingreso registrado y orden creada."
        # --- FIN DE LÓGICA MODIFICADA ---

        # Notificamos al mecánico (solo si no se le notificó al crear la orden)
        if agendamiento.mecanico_asignado and not agendamiento.es_mantenimiento:
            mensaje = f"Se te ha asignado una nueva orden (#{nueva_orden.id}) para el vehículo {nueva_orden.vehiculo.patente}."
            Notificacion.objects.create(
                usuario=agendamiento.mecanico_asignado,
                mensaje=mensaje,
                link=f"/ordenes/{nueva_orden.id}",
            )

            subject_mecanico = f"Nueva Orden Asignada: #{nueva_orden.id}"
            enviar_correo_notificacion(
                agendamiento.mecanico_asignado, subject_mecanico, mensaje
            )

        # Marcamos el agendamiento como FINALIZADO (ya cumplió su propósito)
        agendamiento.estado = Agendamiento.Estado.FINALIZADO
        agendamiento.save()
        
        AgendamientoHistorial.objects.create(
            agendamiento=agendamiento,
            estado=agendamiento.estado, # Guardará "Finalizado"
            usuario=request.user,
            comentario="Vehículo ingresado al taller. Orden creada."
        )

        return Response(
            {
                "message": mensaje_respuesta,
                "orden_id": nueva_orden.id,
            },
            status=status.HTTP_201_CREATED,
        )

    @action(
        detail=True,
        methods=["post"],
        url_path="cancelar",
        permission_classes=[IsSupervisor],
    )
    def cancelar(self, request, pk=None):
        agendamiento = self.get_object()
        agendamiento.estado = Agendamiento.Estado.CANCELADO
        agendamiento.save()
        
        AgendamientoHistorial.objects.create(
            agendamiento=agendamiento,
            estado=agendamiento.estado, # Guardará "Cancelado"
            usuario=request.user,
            comentario="Cita cancelada por el supervisor."
        )
        return Response(
            self.get_serializer(agendamiento).data, status=status.HTTP_200_OK
        )


    @action(
        detail=True,
        methods=["post"],
        url_path="enviar-grua",
        permission_classes=[IsSupervisor],  # Solo supervisores pueden despachar
    )
    def enviar_grua(self, request, pk=None):
        agendamiento = self.get_object()

        if not agendamiento.solicita_grua:
            return Response(
                {"error": "Esta cita no tiene una grúa solicitada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not agendamiento.direccion_grua:
            return Response(
                {"error": "No hay dirección de retiro especificada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if agendamiento.grua_enviada:
            return Response(
                {"error": "La grúa para esta cita ya fue enviada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # 1. Buscamos a todos los usuarios del rol "Grua"
            # (Asegúrate de haber creado este grupo en el Admin o en grupos.json)
            usuarios_grua = User.objects.filter(groups__name="Grua", is_active=True)
            if not usuarios_grua.exists():
                return Response(
                    {"error": "No hay usuarios en el rol 'Grua' para notificar."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # 2. Preparamos el mensaje
            chofer_nombre = agendamiento.chofer_asociado.get_full_name()
            chofer_telefono = agendamiento.chofer_asociado.telefono or "No especificado"
            mensaje = (
                f"NUEVA SOLICITUD DE GRÚA (Cita #{agendamiento.id}):\n"
                f"Vehículo: {agendamiento.vehiculo.patente}\n"
                f"Dirección Retiro: {agendamiento.direccion_grua}\n"
                f"Contacto Chofer: {chofer_nombre} (Tel: {chofer_telefono})"
            )
            subject = f"Solicitud de Grúa - Cita #{agendamiento.id} - Patente {agendamiento.vehiculo.patente}"

            # 3. Enviamos notificación y email a CADA usuario de grúa
            for user_grua in usuarios_grua:
                Notificacion.objects.create(
                    usuario=user_grua,
                    mensaje=f"Nueva solicitud de grúa para {agendamiento.vehiculo.patente}. Dirección: {agendamiento.direccion_grua}",
                    link="/panel-gruas",  # Un futuro panel para ellos
                )
                enviar_correo_notificacion(user_grua, subject, mensaje)

            # 4. Marcamos la grúa como enviada
            agendamiento.grua_enviada = True
            agendamiento.save()

            return Response(
                self.get_serializer(agendamiento).data, status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"error": f"Error inesperado: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class OrdenViewSet(viewsets.ModelViewSet):

    serializer_class = OrdenSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name__in=["Supervisor", "Administrativo"]).exists():
            return (
                Orden.objects.select_related("vehiculo", "usuario_asignado")
                .all()
                .order_by("-fecha_ingreso")
            )
        elif user.groups.filter(name="Mecanico").exists():
            return (
                Orden.objects.filter(usuario_asignado=user)
                .select_related("vehiculo")
                .order_by("-fecha_ingreso")
            )
        elif user.groups.filter(name="Chofer").exists():
            return (
                Orden.objects.filter(vehiculo__chofer=user)
                .select_related("vehiculo")
                .order_by("-fecha_ingreso")
            )
        return Orden.objects.none()

    def get_permissions(self):
        if self.action in ["cambiar_estado"]:
            self.permission_classes = [IsSupervisorOrMecanico]
        return super().get_permissions()

    @action(detail=True, methods=["post"], url_path="cambiar-estado")
    def cambiar_estado(self, request, pk=None):
        orden = self.get_object()
        nuevo_estado = request.data.get("estado")
        motivo = request.data.get("motivo", "")

        try:
            valid_values = list(Orden.Estado.values)
        except Exception:
            valid_values = (
                [c[0] for c in getattr(Orden, "Estado", {}).choices]
                if hasattr(Orden, "Estado")
                else []
            )
        if not nuevo_estado or (valid_values and nuevo_estado not in valid_values):
            return Response(
                {"error": "Debe proporcionar un estado válido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            orden.estado = nuevo_estado

            orden.save()
            OrdenHistorialEstado.objects.create(
                orden=orden, estado=nuevo_estado, usuario=request.user, motivo=motivo
            )

            try:

                chofer_a_notificar = None
                if (
                    orden.agendamiento_origen
                    and orden.agendamiento_origen.chofer_asociado
                ):
                    chofer_a_notificar = orden.agendamiento_origen.chofer_asociado
                elif orden.vehiculo and orden.vehiculo.chofer:  #
                    chofer_a_notificar = orden.vehiculo.chofer

                if chofer_a_notificar:

                    mensajes = {
                        "En Diagnostico": "está siendo diagnosticado por un mecánico.",
                        "En Proceso": "ha entrado en proceso de reparación.",
                        "Pausado": f'ha sido pausado (Motivo: {motivo or "N/A"}).',
                        "Finalizado": "¡está listo! El trabajo en su vehículo ha finalizado.",
                    }

                    mensaje_chofer = mensajes.get(nuevo_estado)
                    if mensaje_chofer:
                        Notificacion.objects.create(
                            usuario=chofer_a_notificar,
                            mensaje=f"Actualización: Su vehículo {orden.vehiculo.patente} {mensaje_chofer}",
                            link="/dashboard",  # El dashboard del chofer
                        )
                        subject_chofer_estado = (
                            f"Actualización Orden #{orden.id}: {orden.vehiculo.patente}"
                        )

                        mensaje_email = f"Actualización: Su vehículo {orden.vehiculo.patente} {mensaje_chofer}"
                        enviar_correo_notificacion(
                            chofer_a_notificar, subject_chofer_estado, mensaje_email
                        )

            except Exception as e:

                print(f"Error al crear notificación de cambio de estado: {e}")

        return Response(self.get_serializer(orden).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="pausar")
    def pausar(self, request, pk=None):
        """Pausa una orden de trabajo."""
        orden = self.get_object()
        motivo = request.data.get("motivo", "Pausa iniciada por el usuario.")

        with transaction.atomic():
            # Cambiamos el estado de la orden a 'Pausado'
            orden.estado = Orden.Estado.PAUSADO
            orden.save()

            # Creamos el registro de la pausa
            OrdenPausa.objects.create(orden=orden, usuario=request.user, motivo=motivo)

            # Guardamos el historial del cambio de estado
            OrdenHistorialEstado.objects.create(
                orden=orden,
                estado=Orden.Estado.PAUSADO,
                usuario=request.user,
                motivo=motivo,
            )

        return Response(self.get_serializer(orden).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="reanudar")
    def reanudar(self, request, pk=None):
        """Reanuda una orden de trabajo que estaba en pausa."""
        orden = self.get_object()

        with transaction.atomic():
            # Buscamos la pausa activa (la que no tiene fecha de fin) y la cerramos
            pausa_activa = orden.pausas.filter(fin__isnull=True).first()
            if pausa_activa:
                pausa_activa.fin = timezone.now()
                pausa_activa.save()

            # Volvemos la orden al estado 'En Proceso' (o el que consideres por defecto)
            orden.estado = Orden.Estado.EN_PROCESO
            orden.save()

            # Guardamos el historial del cambio de estado
            OrdenHistorialEstado.objects.create(
                orden=orden,
                estado=Orden.Estado.EN_PROCESO,
                usuario=request.user,
                motivo="Trabajo reanudado.",
            )

        return Response(self.get_serializer(orden).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="subir-documento")
    def subir_documento(self, request, pk=None):
        """Sube un documento o foto asociado a una orden."""
        orden = self.get_object()
        archivo = request.data.get("archivo")
        if not archivo:
            return Response(
                {"error": "No se envió ningún archivo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Obtener la extensión del archivo, ej: ".pdf" o ".jpg"
        file_name, file_extension = os.path.splitext(archivo.name)
        tipo_detectado = file_extension.lower()  # Guarda la extensión

        # Usamos un serializer específico para la subida de archivos
        serializer = OrdenDocumentoSerializer(
            data=request.data, context={"request": request}
        )
        if serializer.is_valid():
            # Asignamos la orden y el usuario antes de guardar
            serializer.save(
                orden=orden,
                subido_por=request.user,
                estado_en_carga=orden.estado,
                tipo=tipo_detectado,
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class MecanicoListView(generics.ListAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return User.activos.filter(groups__name="Mecanico").order_by("first_name")


class SeguridadAgendaView(generics.ListAPIView):
    serializer_class = AgendamientoSerializer
    permission_classes = [IsSupervisorOrSeguridad]

    def get_queryset(self):
        today = timezone.now().date()
        start = make_aware(datetime.combine(today, datetime.min.time()))
        end = start + timedelta(days=1)
        return Agendamiento.objects.filter(
            estado=Agendamiento.Estado.CONFIRMADO,
            fecha_hora_programada__gte=start,
            fecha_hora_programada__lt=end,
        ).order_by("fecha_hora_programada")


class MisProximasCitasView(generics.ListAPIView):
    """
    Devuelve una lista de las próximas citas con estado 'Confirmado'
    que han sido asignadas al mecánico que realiza la consulta.
    Es una vista de solo lectura para planificación.
    """

    serializer_class = AgendamientoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        # Nos aseguramos de que solo los mecánicos puedan usar esta vista
        if user.groups.filter(name="Mecanico").exists():
            return Agendamiento.objects.filter(
                mecanico_asignado=user,
                estado=Agendamiento.Estado.CONFIRMADO,  # Solo las que no han llegado
            ).order_by("fecha_hora_programada")

        # Si no es mecánico, no devolvemos nada
        return Agendamiento.objects.none()


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mecanico_dashboard_stats(request):
    """
    Prepara y devuelve las estadísticas y tareas para el dashboard del mecánico.
    """
    user = request.user
    if not user.groups.filter(name="Mecanico").exists():
        return Response(
            {"error": "Acceso no autorizado"}, status=status.HTTP_403_FORBIDDEN
        )

    # 1. Contar órdenes activas (todas las que no estén 'Finalizado')
    ordenes_activas_count = (
        Orden.objects.filter(usuario_asignado=user)
        .exclude(estado=Orden.Estado.FINALIZADO)
        .count()
    )

    # 2. Contar próximas asignaciones (agendamientos confirmados pero sin orden creada)
    proximas_asignaciones_count = Agendamiento.objects.filter(
        mecanico_asignado=user, estado=Agendamiento.Estado.CONFIRMADO
    ).count()

    # 3. Obtener la lista de órdenes activas
    ordenes_activas = (
        Orden.objects.filter(usuario_asignado=user)
        .exclude(estado=Orden.Estado.FINALIZADO)
        .order_by("fecha_ingreso")
    )

    # Serializar la lista de órdenes
    ordenes_serializer = OrdenSerializer(
        ordenes_activas, many=True, context={"request": request}
    )

    # Construir la respuesta
    data = {
        "kpis": {
            "ordenesActivas": ordenes_activas_count,
            "proximasAsignaciones": proximas_asignaciones_count,
        },
        "tareas": ordenes_serializer.data,
    }
    return Response(data, status=status.HTTP_200_OK)


class OrdenesPendientesSalidaView(APIView):
    """
    Endpoint [GET] para listar todas las órdenes de trabajo que
    están en estado 'Finalizado' pero aún no tienen una
    fecha de entrega real (es decir, no han salido del taller).
    """

    def get(self, request, *args, **kwargs):
        try:

            ordenes_listas = (
                Orden.objects.filter(
                    estado=Orden.Estado.FINALIZADO, fecha_entrega_real__isnull=True
                )
                .select_related(
                    "vehiculo",
                    "agendamiento_origen__chofer_asociado",
                    "usuario_asignado",
                )
                .order_by("fecha_ingreso")
            )

            serializer = OrdenSalidaListSerializer(ordenes_listas, many=True)

            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error al obtener órdenes: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class RegistrarSalidaView(APIView):
    """
    Endpoint [POST] para registrar la salida de un vehículo.
    Recibe el ID (pk) de la Orden en la URL.
    """

    def post(self, request, pk, *args, **kwargs):
        try:
            # 1. Buscamos la orden
            orden = get_object_or_404(Orden, pk=pk)

            # 2. Validaciones
            if orden.fecha_entrega_real:
                return Response(
                    {"error": "Esta salida ya fue registrada."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if orden.estado != Orden.Estado.FINALIZADO:
                return Response(
                    {
                        "error": "El trabajo en este vehículo aún no ha sido finalizado por el taller."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            orden.fecha_entrega_real = timezone.now()

            if orden.agendamiento_origen:
                if orden.agendamiento_origen.estado != Agendamiento.Estado.FINALIZADO:
                    orden.agendamiento_origen.estado = Agendamiento.Estado.FINALIZADO
                    orden.agendamiento_origen.save()

            orden.save()

            return Response(
                {
                    "mensaje": f"Salida del vehículo {orden.vehiculo.patente} registrada con éxito."
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": f"Error al registrar la salida: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class MecanicoAgendaView(generics.ListAPIView):
    serializer_class = AgendamientoSerializer
    permission_classes = [IsSupervisor]

    def get_queryset(self):
        mecanico_id = self.kwargs.get("mecanico_id")
        if not mecanico_id:
            return Agendamiento.objects.none()

        queryset = Agendamiento.objects.filter(
            mecanico_asignado_id=mecanico_id,
            estado__in=[Agendamiento.Estado.CONFIRMADO, Agendamiento.Estado.EN_TALLER],
        )

        fecha_str = self.request.query_params.get("fecha")
        if fecha_str:
            try:
                fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()

                current_tz = timezone.get_current_timezone()

                start_of_day = datetime.combine(
                    fecha, datetime.min.time(), tzinfo=current_tz
                )
                end_of_day = datetime.combine(
                    fecha, datetime.max.time(), tzinfo=current_tz
                )

                queryset = queryset.filter(
                    fecha_hora_programada__gte=start_of_day,
                    fecha_hora_programada__lte=end_of_day,
                )
            except ValueError:
                pass

        return queryset.order_by("fecha_hora_programada")


class NotificacionViewSet(viewsets.ModelViewSet):
    """
    API para leer, crear y marcar notificaciones como leídas.
    """

    serializer_class = NotificacionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtra notificaciones solo para el usuario logueado."""
        return Notificacion.objects.filter(usuario=self.request.user).order_by("-fecha")

    @action(detail=False, methods=["post"], url_path="marcar-como-leidas")
    def marcar_como_leidas(self, request):
        """Acción para marcar todas las notificaciones del usuario como leídas."""
        Notificacion.objects.filter(usuario=request.user, leida=False).update(
            leida=True
        )
        return Response(status=status.HTTP_204_NO_CONTENT)


class TallerViewSet(viewsets.ModelViewSet):
    """
    API para gestionar los Talleres.
    Solo Supervisores/Admin pueden crear o editar talleres.
    """

    queryset = Taller.objects.all()
    serializer_class = TallerSerializer
    permission_classes = [IsSupervisor]  # Protegido

    def get_permissions(self):
        if self.action in ["list", "retrieve"]:
            # Permite a cualquier usuario autenticado (ej. Mecánico) ver la lista
            self.permission_classes = [IsAuthenticated]
        return super().get_permissions()


class ProductoViewSet(viewsets.ModelViewSet):
    """
    API para buscar productos (repuestos) y ver stock.
    Usado por Mecánicos y Repuestos.
    """

    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [IsAuthenticated]  # Abierto a todos los logueados
    filter_backends = [filters.SearchFilter]
    search_fields = ["nombre", "sku", "marca"]
    lookup_field = "sku"

    def create(self, request, *args, **kwargs):
        # Copiamos los datos para poder modificarlos
        data = request.data.copy()

        # Verificamos si es una creación (sin SKU) y si hay un nombre
        if "sku" not in data or not data["sku"]:
            nombre = data.get("nombre")
            if not nombre:
                return Response(
                    {"error": "El campo 'nombre' es obligatorio para crear un SKU."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # --- Lógica de generación de SKU ---

            # 1. Generar prefijo (Ej: "Frenos" -> "FRE")
            prefix = nombre[:3].upper()
            base_sku = f"{prefix}-"

            # 2. Buscar el último número usado para este prefijo
            # Usamos .order_by('sku').last() para obtener el más alto alfanuméricamente
            last_product = (
                Producto.objects.filter(sku__startswith=base_sku).order_by("sku").last()
            )

            next_num = 101  # Empezamos en 101 si es el primero

            if last_product:
                try:
                    # Extraer el número del último SKU (Ej: "FRE-273" -> "273")
                    last_num_str = last_product.sku.split("-")[-1]
                    next_num = int(last_num_str) + 1
                except (ValueError, IndexError):
                    # Si falla (ej: SKU es "FRE-ABC"), usamos el 'next_num' por defecto
                    pass

            # 3. Bucle de seguridad para garantizar unicidad (evita "race conditions")
            while True:
                new_sku = f"{base_sku}{next_num}"
                if not Producto.objects.filter(sku=new_sku).exists():
                    data["sku"] = new_sku  # Asignamos el SKU único
                    break
                next_num += 1  # Si "FRE-101" existe, prueba "FRE-102"

            # --- Fin de la lógica ---

        # 4. Continuar con la creación normal usando los datos modificados
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )


class OrdenItemViewSet(viewsets.ModelViewSet):
    """
    API para gestionar los items (repuestos) de una orden.
    """

    queryset = OrdenItem.objects.all()
    serializer_class = OrdenItemSerializer
    permission_classes = [IsAuthenticated]  # Permisos más granulares por acción

    def get_queryset(self):
        """
        Filtra para que cada rol vea lo que le corresponde.
        """
        user = self.request.user
        if user.groups.filter(
            name__in=["Supervisor", "Administrativo", "Repuestos"]
        ).exists():
            return OrdenItem.objects.all()
        if user.groups.filter(name="Mecanico").exists():
            return OrdenItem.objects.filter(solicitado_por=user)
        # Choferes u otros no deberían ver esto directamente
        return OrdenItem.objects.none()

    def get_permissions(self):
        """Permisos por acción"""
        if self.action in ["aprobar_repuesto", "rechazar_repuesto", "list_pendientes"]:
            self.permission_classes = [IsRepuestos]
        elif self.action in ["create", "list", "retrieve"]:
            self.permission_classes = [IsSupervisorOrMecanico]
        return super().get_permissions()

    def perform_create(self, serializer):
        """
        _Asigna al mecánico que solicita y notifica a Repuestos._
        """
        item = serializer.save(solicitado_por=self.request.user)

        # Si es un producto, notificar a Repuestos
        if item.producto:
            try:
                usuarios_repuestos = User.objects.filter(
                    groups__name__in=["Repuestos", "Supervisor"], is_active=True
                )
                mecanico_nombre = (
                    self.request.user.first_name or self.request.user.username
                )
                mensaje = (
                    f"El mecánico {mecanico_nombre} solicitó {item.cantidad}x {item.producto.nombre} "
                    f"para la Orden #{item.orden.id}."
                )
                subject = f"Nueva Solicitud de Repuesto: Orden #{item.orden.id}"

                for user_rep in usuarios_repuestos:
                    Notificacion.objects.create(
                        usuario=user_rep, mensaje=mensaje, link="/panel-repuestos"
                    )  # Link a la nueva página de repuestos
                    enviar_correo_notificacion(
                        user_rep, subject, mensaje
                    )  # Opcional si quieres email
            except Exception as e:
                print(f"ERROR al notificar a Repuestos: {e}")

    @action(detail=False, methods=["get"], url_path="pendientes")
    def list_pendientes(self, request):
        """
        _Endpoint para que el rol Repuestos vea solo lo pendiente._
        """
        pendientes = OrdenItem.objects.filter(
            producto__isnull=False, estado_repuesto=OrdenItem.EstadoRepuesto.PENDIENTE
        ).select_related("orden", "producto", "solicitado_por")

        serializer = self.get_serializer(pendientes, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="gestionar-repuesto")
    def gestionar_repuesto(self, request, pk=None):
        """
        _Endpoint para que Repuestos APRUEBE o RECHACE._
        """
        item = self.get_object()
        accion = request.data.get("accion")  # "aprobar" o "rechazar"
        motivo = request.data.get("motivo", "")

        if not accion or accion not in ["aprobar", "rechazar"]:
            return Response(
                {"error": "Acción no válida."}, status=status.HTTP_400_BAD_REQUEST
            )

        if item.estado_repuesto != OrdenItem.EstadoRepuesto.PENDIENTE:
            return Response(
                {"error": "Esta solicitud ya fue gestionada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            if accion == "aprobar":
                if item.producto.stock < item.cantidad:
                    return Response(
                        {
                            "error": f"Stock insuficiente. Solo quedan {item.producto.stock}."
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Descontar stock
                item.producto.stock = F("stock") - item.cantidad
                item.producto.save()
                item.producto.refresh_from_db()

                # Actualizar item
                item.estado_repuesto = OrdenItem.EstadoRepuesto.APROBADO
                item.gestionado_por = request.user
                item.fecha_gestion = timezone.now()
                item.save()
                item.refresh_from_db()

                # Notificar al mecánico
                subject_mec = f"Repuesto Aprobado: Orden #{item.orden.id}"
                mensaje_mec = f"Su solicitud de {item.cantidad}x {item.producto.nombre} fue APROBADA."
                Notificacion.objects.create(
                    usuario=item.solicitado_por,
                    mensaje=mensaje_mec,
                    link=f"/ordenes/{item.orden.id}",
                )
                enviar_correo_notificacion(
                    item.solicitado_por, subject_mec, mensaje_mec
                )

            elif accion == "rechazar":
                item.estado_repuesto = OrdenItem.EstadoRepuesto.RECHAZADO
                item.gestionado_por = request.user
                item.fecha_gestion = timezone.now()
                item.motivo_gestion = (
                    motivo or "Sin stock. Solicitado a proveedor (3 días aprox.)"
                )
                item.save()
                item.refresh_from_db()

                # Notificar al mecánico
                subject_mec = f"Repuesto Rechazado: Orden #{item.orden.id}"
                mensaje_mec = f"Su solicitud de {item.cantidad}x {item.producto.nombre} fue RECHAZADA. Motivo: {item.motivo_gestion}"
                Notificacion.objects.create(
                    usuario=item.solicitado_por,
                    mensaje=mensaje_mec,
                    link=f"/ordenes/{item.orden.id}",
                )
                enviar_correo_notificacion(
                    item.solicitado_por, subject_mec, mensaje_mec
                )

        return Response(self.get_serializer(item).data, status=status.HTTP_200_OK)


# ======================================================================
# 🔑 API DE GESTIÓN DE LLAVES
# ======================================================================


class LlaveVehiculoViewSet(viewsets.ModelViewSet):
    """
    API para gestionar el inventario de llaves (Pañol).
    Cubre: Control de duplicados, Reportar pérdidas.
    """

    queryset = LlaveVehiculo.objects.all().select_related("vehiculo", "poseedor_actual")
    serializer_class = LlaveVehiculoSerializer
    permission_classes = [IsControlLlaves]  # Protegido para el nuevo rol

    @action(detail=True, methods=["post"], url_path="registrar-devolucion")
    @transaction.atomic
    def registrar_devolucion(self, request, pk=None):
        """
        Cierra un préstamo activo (RECIBIR LLAVE).
        El 'pk' aquí es el ID de la *LlaveVehiculo*.
        """
        llave = self.get_object()

        try:

            prestamo = PrestamoLlave.objects.get(
                llave=llave, fecha_hora_devolucion__isnull=True
            )
        except PrestamoLlave.DoesNotExist:
            return Response(
                {"error": "Esta llave no tiene un préstamo activo para devolver."},
                status=status.HTTP_404_NOT_FOUND,
            )

        prestamo.fecha_hora_devolucion = timezone.now()
        prestamo.observaciones_devolucion = request.data.get("observaciones", "")
        prestamo.save()

        llave.estado = LlaveVehiculo.Estado.EN_BODEGA
        llave.poseedor_actual = None
        llave.save()

        serializer = self.get_serializer(llave)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="registrar-retiro")
    @transaction.atomic
    def registrar_retiro(self, request, pk=None):
        """
        Crea un nuevo préstamo (PRESTAR LLAVE). El 'pk' es el ID de la Llave.
        """
        llave = self.get_object()
        usuario_id = request.data.get("usuario_id")
        observaciones = request.data.get("observaciones", "")

        if llave.estado != LlaveVehiculo.Estado.EN_BODEGA:
            return Response(
                {"error": "La llave no existe o ya está prestada."},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            usuario = User.objects.get(id=usuario_id, is_active=True)
        except User.DoesNotExist:
            return Response(
                {"error": "El usuario seleccionado no existe."},
                status=status.HTTP_404_NOT_FOUND,
            )

        prestamo = PrestamoLlave.objects.create(
            llave=llave, usuario_retira=usuario, observaciones_retiro=observaciones
        )

        llave.estado = LlaveVehiculo.Estado.PRESTADA
        llave.poseedor_actual = usuario
        llave.save()

        serializer = self.get_serializer(llave)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"], url_path="reportar-estado")
    @transaction.atomic
    def reportar_estado(self, request, pk=None):
        """
        Acción para marcar una llave como 'Perdida' o 'Dañada'
        """
        llave = self.get_object()
        nuevo_estado = request.data.get("estado")  # "Perdida" o "Dañada"
        motivo = request.data.get("motivo")

        if not motivo:
            return Response(
                {"error": "Se requiere un motivo para el reporte."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if nuevo_estado not in [
            LlaveVehiculo.Estado.PERDIDA,
            LlaveVehiculo.Estado.DAÑADA,
        ]:
            return Response(
                {
                    "error": 'Estado no válido. Solo se puede reportar como "Perdida" o "Dañada".'
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if llave.estado != LlaveVehiculo.Estado.EN_BODEGA:
            return Response(
                {
                    "error": 'Solo se pueden reportar llaves que están "En Bodega". Devuelva la llave antes de reportarla.'
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        estado_anterior = llave.estado
        llave.estado = nuevo_estado
        llave.motivo_reporte = motivo
        llave.save()

        LlaveHistorialEstado.objects.create(
            llave=llave,
            usuario_reporta=request.user,
            estado_anterior=estado_anterior,
            estado_nuevo=nuevo_estado,
            motivo=motivo,
        )

        serializer = self.get_serializer(llave)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="revertir-reporte")
    @transaction.atomic
    def revertir_reporte(self, request, pk=None):
        """
        Revierte un reporte, marcando la llave como "En Bodega"
        y limpiando el motivo.
        """
        llave = self.get_object()

        if llave.estado not in [
            LlaveVehiculo.Estado.PERDIDA,
            LlaveVehiculo.Estado.DAÑADA,
        ]:
            return Response(
                {"error": "Esta llave no tiene un reporte activo que revertir."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        estado_anterior = llave.estado  # <-- Captura el estado actual
        llave.estado = LlaveVehiculo.Estado.EN_BODEGA
        llave.motivo_reporte = None  # Limpiamos el motivo
        llave.save()

        LlaveHistorialEstado.objects.create(
            llave=llave,
            usuario_reporta=request.user,
            estado_anterior=estado_anterior,
            estado_nuevo=LlaveVehiculo.Estado.EN_BODEGA,
            motivo="Reporte revertido. Llave vuelve a estar operativa.",
        )

        serializer = self.get_serializer(llave)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="historial")
    def historial(self, request, pk=None):
        """
        Devuelve el historial de préstamos de una llave específica.
        """
        llave = self.get_object()
        prestamos = llave.prestamos.all().select_related("usuario_retira")
        serializer = PrestamoLlaveSerializer(prestamos, many=True)
        return Response(serializer.data)


class PrestamoLlaveViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API para LEER el historial de préstamos.
    """

    queryset = PrestamoLlave.objects.all().select_related(
        "llave__vehiculo", "usuario_retira"
    )
    serializer_class = PrestamoLlaveSerializer
    permission_classes = [IsControlLlaves]

    filter_backends = [filters.SearchFilter]
    search_fields = [
        "llave__vehiculo__patente",
        "usuario_retira__username",
        "usuario_retira__first_name",
        "usuario_retira__last_name",
    ]

    @action(detail=True, methods=["post"], url_path="registrar-devolucion")
    @transaction.atomic
    def registrar_devolucion(self, request, pk=None):
        """
        Cierra un préstamo activo (RECIBIR LLAVE).
        El 'pk' aquí es el ID de la *LlaveVehiculo* para encontrar el préstamo activo.
        """
        try:

            prestamo = PrestamoLlave.objects.get(
                llave_id=pk, fecha_hora_devolucion__isnull=True
            )
        except PrestamoLlave.DoesNotExist:
            return Response(
                {"error": "Esta llave no tiene un préstamo activo para devolver."},
                status=status.HTTP_404_NOT_FOUND,
            )

        prestamo.fecha_hora_devolucion = timezone.now()
        prestamo.observaciones_devolucion = request.data.get("observaciones", "")
        prestamo.save()

        llave = prestamo.llave
        llave.estado = LlaveVehiculo.Estado.EN_BODEGA
        llave.poseedor_actual = None
        llave.save()

        serializer = self.get_serializer(prestamo)
        return Response(serializer.data, status=status.HTTP_200_OK)


class LlaveHistorialEstadoViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API para LEER el historial de reportes de llaves.
    """

    queryset = LlaveHistorialEstado.objects.all().select_related(
        "llave__vehiculo", "usuario_reporta"
    )
    serializer_class = LlaveHistorialEstadoSerializer
    permission_classes = [IsControlLlaves]

    filter_backends = [filters.SearchFilter]

    search_fields = [
        "llave__vehiculo__patente",
        "usuario_reporta__username",
        "usuario_reporta__first_name",
        "usuario_reporta__last_name",
    ]


class HistorialSeguridadViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API de solo lectura para el historial de ingresos y salidas de Seguridad.
    """

    serializer_class = HistorialSeguridadSerializer
    permission_classes = [IsSupervisorOrSeguridad]

    filter_backends = [filters.SearchFilter]

    search_fields = [
        "vehiculo__patente",
        "agendamiento_origen__chofer_asociado__first_name",
        "agendamiento_origen__chofer_asociado__last_name",
        "vehiculo__chofer__first_name",
        "vehiculo__chofer__last_name",
    ]

    def get_queryset(self):
        return (
            Orden.objects.all()
            .select_related(
                "vehiculo", "agendamiento_origen__chofer_asociado", "vehiculo__chofer"
            )
            .order_by("-fecha_ingreso")
        )


# Todos los reportes


@api_view(["GET"])
@permission_classes([IsAdministrativo])  # Protegemos la vista
def exportar_bitacora_seguridad(request):
    """
    Genera un reporte Excel de la bitácora de Ingresos y Salidas.
    Acepta filtros de fecha: ?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD
    """

    # 1. Obtener filtros de fecha de la URL
    fecha_inicio_str = request.query_params.get("fecha_inicio")
    fecha_fin_str = request.query_params.get("fecha_fin")

    # 2. Query base
    queryset = (
        Orden.objects.all()
        .select_related(
            "vehiculo", "agendamiento_origen__chofer_asociado", "vehiculo__chofer"
        )
        .order_by("fecha_ingreso")
    )

    # 3. Aplicar filtros de fecha si existen
    if fecha_inicio_str and fecha_fin_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
            # Combinamos la fecha fin con la última hora del día para incluir todo el día
            fecha_fin = datetime.combine(
                datetime.strptime(fecha_fin_str, "%Y-%m-%d").date(), time.max
            )
            # Filtramos por el rango de fecha de INGRESO
            queryset = queryset.filter(fecha_ingreso__range=[fecha_inicio, fecha_fin])
        except ValueError:
            return Response(
                {"error": "Formato de fecha inválido. Usar YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    # 4. Crear el libro de Excel en memoria
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bitácora de Movimientos"

    # 5. Definir Títulos de las columnas
    columnas = [
        "ID Orden",
        "Patente",
        "Chofer",
        "Fecha Ingreso",
        "Fecha Salida",
        "Estado Actual",
    ]
    ws.append(columnas)

    # Añadir estilo simple a la cabecera (Negrita)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # 6. Llenar el Excel con los datos
    for orden in queryset:

        # Lógica para obtener el nombre del chofer (como en tu Serializer)
        chofer_nombre = "No asignado"
        if orden.agendamiento_origen and orden.agendamiento_origen.chofer_asociado:
            chofer_nombre = orden.agendamiento_origen.chofer_asociado.get_full_name()
        elif orden.vehiculo and orden.vehiculo.chofer:
            chofer_nombre = orden.vehiculo.chofer.get_full_name()

        # Formatear fechas para que Excel las entienda
        fecha_ingreso_excel = (
            orden.fecha_ingreso.replace(tzinfo=None) if orden.fecha_ingreso else None
        )
        fecha_salida_excel = (
            orden.fecha_entrega_real.replace(tzinfo=None)
            if orden.fecha_entrega_real
            else "En Taller"
        )

        fila = [
            orden.id,
            orden.vehiculo.patente if orden.vehiculo else "S/P",
            chofer_nombre,
            fecha_ingreso_excel,
            fecha_salida_excel,
            orden.estado,
        ]
        ws.append(fila)

        # Aplicar formato de fecha a las celdas
        if fecha_ingreso_excel:
            ws.cell(row=ws.max_row, column=4).number_format = "DD/MM/YYYY HH:MM"
        if isinstance(fecha_salida_excel, datetime):
            ws.cell(row=ws.max_row, column=5).number_format = "DD/MM/YYYY HH:MM"

    # 7. Crear la respuesta HTTP con el archivo
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # Definir el nombre del archivo
    response["Content-Disposition"] = (
        f'attachment; filename="Reporte_Seguridad_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )

    # Guardar el libro de Excel en la respuesta
    wb.save(response)

    return response


@api_view(["GET"])
@permission_classes([IsAdministrativo])  # Reutilizamos el mismo permiso
def exportar_snapshot_taller_pdf(request):
    """
    Genera un reporte PDF (Snapshot) de los vehículos
    actualmente en el taller.
    """

    # 1. Query: Vehículos que NO están 'Finalizados'
    ordenes_activas = (
        Orden.objects.exclude(estado=Orden.Estado.FINALIZADO)
        .select_related(
            "vehiculo", "agendamiento_origen__chofer_asociado", "vehiculo__chofer"
        )
        .order_by("fecha_ingreso")
    )

    # 2. Preparar el PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()

    # 3. Títulos del PDF
    fecha_actual = timezone.now().strftime("%d/%m/%Y %H:%M:%S")
    elements.append(Paragraph("Reporte de Vehículos en Taller", styles["h1"]))
    elements.append(Paragraph(f"Generado el: {fecha_actual}", styles["Normal"]))
    elements.append(
        Paragraph(f"Total Vehículos: {ordenes_activas.count()}", styles["Normal"])
    )
    elements.append(Paragraph(" ", styles["Normal"]))  # Espacio

    # 4. Preparar datos para la tabla
    data = [["Patente", "Chofer", "Fecha Ingreso", "Estado Actual"]]

    for orden in ordenes_activas:
        # Lógica para obtener el nombre del chofer
        chofer_nombre = "No asignado"
        if orden.agendamiento_origen and orden.agendamiento_origen.chofer_asociado:
            chofer_nombre = orden.agendamiento_origen.chofer_asociado.get_full_name()
        elif orden.vehiculo and orden.vehiculo.chofer:
            chofer_nombre = orden.vehiculo.chofer.get_full_name()

        fecha_ingreso_str = (
            orden.fecha_ingreso.strftime("%d/%m/%Y %H:%M")
            if orden.fecha_ingreso
            else "N/A"
        )

        data.append(
            [
                orden.vehiculo.patente if orden.vehiculo else "S/P",
                chofer_nombre,
                fecha_ingreso_str,
                orden.estado,
            ]
        )

    # 5. Crear y Estilizar la Tabla
    table = Table(data)
    style = TableStyle(
        [
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#2d3748"),
            ),  # Fondo cabecera
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#4a5568")),  # Fondo filas
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]
    )
    table.setStyle(style)
    elements.append(table)

    # 6. Construir el PDF
    doc.build(elements)

    # 7. Crear la respuesta HTTP
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    filename = f"Snapshot_Taller_{timezone.now().strftime('%Y%m%d')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@api_view(["GET"])
@permission_classes([IsAdministrativo])  # Reutilizamos el permiso
def exportar_consumo_repuestos(request):
    """
    Genera un reporte Excel del consumo de repuestos aprobados.
    Acepta filtros de fecha: ?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD
    """

    # 1. Obtener filtros de fecha de la URL
    fecha_inicio_str = request.query_params.get("fecha_inicio")
    fecha_fin_str = request.query_params.get("fecha_fin")

    # 2. Query base: Filtramos solo repuestos 'Aprobados'
    queryset = (
        OrdenItem.objects.filter(estado_repuesto=OrdenItem.EstadoRepuesto.APROBADO)
        .select_related(
            "orden", "producto", "solicitado_por"  # El mecánico que pidió el repuesto
        )
        .order_by("fecha_gestion")
    )  # Ordenamos por la fecha en que se aprobó

    # 3. Aplicar filtros de fecha (sobre la fecha de gestión/aprobación)
    if fecha_inicio_str and fecha_fin_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
            fecha_fin = datetime.combine(
                datetime.strptime(fecha_fin_str, "%Y-%m-%d").date(), time.max
            )
            # Filtramos por el rango de fecha de APROBACIÓN
            queryset = queryset.filter(fecha_gestion__range=[fecha_inicio, fecha_fin])
        except ValueError:
            return Response(
                {"error": "Formato de fecha inválido. Usar YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    # 4. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumo Repuestos"

    # 5. Definir Títulos de las columnas
    columnas = [
        "Fecha Aprobado",
        "ID Orden",
        "Mecánico Solicitante",
        "SKU",
        "Producto",
        "Cantidad",
        "Precio Unitario",
        "Costo Total",
    ]
    ws.append(columnas)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # 6. Llenar el Excel con los datos
    total_general = 0
    for item in queryset:

        # Calculamos el costo total del ítem
        costo_total = item.cantidad * item.precio_unitario
        total_general += costo_total

        fecha_gestion_excel = (
            item.fecha_gestion.replace(tzinfo=None) if item.fecha_gestion else None
        )

        fila = [
            fecha_gestion_excel,
            item.orden.id if item.orden else "N/A",
            item.solicitado_por.get_full_name() if item.solicitado_por else "N/A",
            item.producto.sku if item.producto else "N/A",
            item.producto.nombre if item.producto else "N/A",
            item.cantidad,
            item.precio_unitario,
            costo_total,
        ]
        ws.append(fila)

        # Aplicar formato de fecha y moneda
        ws.cell(row=ws.max_row, column=1).number_format = "DD/MM/YYYY HH:MM"
        ws.cell(row=ws.max_row, column=7).number_format = "$ #,##0"
        ws.cell(row=ws.max_row, column=8).number_format = "$ #,##0"

    # Añadir fila de Total General
    ws.append([])  # Fila vacía
    ws.append([None, None, None, None, None, None, "Total General:", total_general])
    ws.cell(row=ws.max_row, column=7).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=ws.max_row, column=8).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=ws.max_row, column=8).number_format = "$ #,##0"

    # 7. Crear la respuesta HTTP con el archivo
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Reporte_Consumo_Repuestos_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)

    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdministrativo])
def exportar_inventario_valorizado(request):
    """
    Exporta un snapshot del inventario actual y su valor.
    No requiere filtros de fecha.
    """

    # 1. Obtener todos los productos
    productos = Producto.objects.all().order_by("nombre")

    # 2. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Valorizado"

    # 3. Definir las cabeceras
    headers = [
        "SKU",
        "Nombre Producto",
        "Marca",
        "Stock Actual",
        "Precio Unitario",
        "Valor Total Inventario",
    ]
    ws.append(headers)

    # 4. Añadir los datos
    for producto in productos:
        valor_total = producto.stock * producto.precio_venta

        ws.append(
            [
                producto.sku,
                producto.nombre,
                producto.marca,
                producto.stock,
                producto.precio_venta,
                valor_total,
            ]
        )

    # 5. Preparar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # Formateamos el nombre del archivo con la fecha/hora actual
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = (
        f'attachment; filename="Reporte_Inventario_Valorizado_{timestamp}.xlsx"'
    )

    # 6. Guardar el libro en la respuesta
    wb.save(response)
    return response


# NUEVA VISTA 2: REPORTE DE QUIEBRES DE STOCK (CON FILTROS DE FECHA)
@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdministrativo])
def exportar_quiebres_stock(request):
    """
    Exporta un reporte de todos los repuestos solicitados que fueron RECHAZADOS.
    Utiliza los filtros de fecha (fecha_inicio, fecha_fin) sobre 'fecha_gestion'.
    """

    # 1. Obtener filtros de fecha (igual que en tus otras vistas)
    fecha_inicio_str = request.query_params.get("fecha_inicio", None)
    fecha_fin_str = request.query_params.get("fecha_fin", None)

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Error: Faltan filtros de fecha.", status=400)

    # Convertimos strings a objetos datetime
    # (Ajusta el formato si es necesario, pero 'YYYY-MM-DD' es estándar)
    fecha_inicio_dt = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
    fecha_fin_dt = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()

    # 2. Obtener los items rechazados en ese rango
    items_rechazados = (
        OrdenItem.objects.filter(
            estado_repuesto=OrdenItem.EstadoRepuesto.RECHAZADO,
            fecha_gestion__range=[fecha_inicio_dt, fecha_fin_dt],
        )
        .select_related("producto", "solicitado_por", "orden")
        .order_by("-fecha_gestion")
    )

    # 3. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiebres de Stock"

    # 4. Definir las cabeceras
    headers = [
        "Fecha Rechazo",
        "Producto",
        "SKU",
        "Cantidad Solicitada",
        "Solicitado Por (Mecánico)",
        "ID Orden",
        "Motivo del Rechazo",
    ]
    ws.append(headers)

    # 5. Añadir los datos
    for item in items_rechazados:
        mecanico_nombre = (
            item.solicitado_por.get_full_name() if item.solicitado_por else "N/A"
        )

        ws.append(
            [
                item.fecha_gestion,
                item.producto.nombre,
                item.producto.sku,
                item.cantidad,
                mecanico_nombre,
                item.orden.id,
                item.motivo_gestion,
            ]
        )

    # 6. Preparar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Reporte_Quiebres_Stock_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"'
    )
    wb.save(response)
    return response


# NUEVA VISTA 1: REPORTE DE PRODUCTIVIDAD POR MECÁNICO
@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdministrativo])
def exportar_productividad_mecanicos(request):
    """
    Exporta un reporte de productividad (Órdenes Finalizadas)
    agrupado por mecánico, dentro de un rango de fechas.
    """

    # 1. Obtener filtros de fecha
    fecha_inicio_str = request.query_params.get("fecha_inicio", None)
    fecha_fin_str = request.query_params.get("fecha_fin", None)

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Error: Faltan filtros de fecha.", status=400)

    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        # Ajustamos fecha_fin para incluir el día completo
        fecha_fin_dt = datetime.combine(
            datetime.strptime(fecha_fin_str, "%Y-%m-%d").date(), time.max
        )
    except ValueError:
        return HttpResponse("Error: Formato de fecha inválido.", status=400)

    # 2. Query: Agrupar por mecánico y contar órdenes finalizadas
    productividad = (
        Orden.objects.filter(
            estado=Orden.Estado.FINALIZADO,
            fecha_entrega_real__range=[fecha_inicio_dt, fecha_fin_dt],
            usuario_asignado__isnull=False,  # Asegurarnos de que tenga un mecánico
        )
        .values(
            "usuario_asignado__first_name",  # Agrupar por nombre
            "usuario_asignado__last_name",  # Agrupar por apellido
            "usuario_asignado__rut",  # Agrupar por RUT (para ID único)
        )
        .annotate(ordenes_finalizadas=Count("id"))  # Contar las órdenes para ese grupo
        .order_by("-ordenes_finalizadas")
    )  # Ordenar de más productivo a menos

    # 3. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productividad Mecánicos"

    # 4. Definir las cabeceras
    headers = ["RUT Mecánico", "Nombre Mecánico", "Órdenes Finalizadas"]
    ws.append(headers)

    # 5. Añadir los datos
    for data in productividad:
        nombre_completo = f"{data['usuario_asignado__first_name']} {data['usuario_asignado__last_name']}"
        ws.append(
            [
                data["usuario_asignado__rut"],
                nombre_completo,
                data["ordenes_finalizadas"],
            ]
        )

    # 6. Preparar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Reporte_Productividad_Mecanicos_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"'
    )
    wb.save(response)
    return response


# NUEVA VISTA 2: REPORTE DE TIEMPOS DE TALLER
@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdministrativo])
def exportar_tiempos_taller(request):
    """
    Exporta un reporte detallado de los tiempos por Orden.
    Tiempo total en taller vs Tiempo total en Pausa.
    """

    # 1. Obtener filtros de fecha (sobre órdenes FINALIZADAS en ese rango)
    fecha_inicio_str = request.query_params.get("fecha_inicio", None)
    fecha_fin_str = request.query_params.get("fecha_fin", None)

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Error: Faltan filtros de fecha.", status=400)

    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        fecha_fin_dt = datetime.combine(
            datetime.strptime(fecha_fin_str, "%Y-%m-%d").date(), time.max
        )
    except ValueError:
        return HttpResponse("Error: Formato de fecha inválido.", status=400)

    # 2. Query: Órdenes finalizadas en el rango
    ordenes = (
        Orden.objects.filter(
            estado=Orden.Estado.FINALIZADO,
            fecha_entrega_real__range=[fecha_inicio_dt, fecha_fin_dt],
        )
        .select_related("usuario_asignado", "vehiculo")
        .prefetch_related("pausas")  # prefetch_related para las pausas (muchos a uno)
        .order_by("fecha_entrega_real")
    )

    # 3. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tiempos de Taller"

    # 4. Definir las cabeceras
    headers = [
        "ID Orden",
        "Patente",
        "Mecánico",
        "Fecha Ingreso",
        "Fecha Salida",
        "Tiempo Total Taller (Horas)",
        "Tiempo Total Pausas (Horas)",
        "Tiempo Efectivo (Horas)",
    ]
    ws.append(headers)

    # 5. Añadir los datos (Calculamos los tiempos)
    for orden in ordenes:

        # Cálculo 1: Tiempo Total en Taller
        tiempo_total_taller_horas = 0
        if orden.fecha_ingreso and orden.fecha_entrega_real:
            duracion_total = orden.fecha_entrega_real - orden.fecha_ingreso
            tiempo_total_taller_horas = round(duracion_total.total_seconds() / 3600, 2)

        # Cálculo 2: Tiempo Total en Pausas
        tiempo_total_pausas_horas = 0
        pausas = orden.pausas.all()
        for pausa in pausas:
            if pausa.fin and pausa.inicio:
                duracion_pausa = pausa.fin - pausa.inicio
                tiempo_total_pausas_horas += round(
                    duracion_pausa.total_seconds() / 3600, 2
                )

        # Cálculo 3: Tiempo Efectivo
        tiempo_efectivo_horas = round(
            tiempo_total_taller_horas - tiempo_total_pausas_horas, 2
        )

        mecanico_nombre = (
            orden.usuario_asignado.get_full_name() if orden.usuario_asignado else "N/A"
        )

        ws.append(
            [
                orden.id,
                orden.vehiculo.patente,
                mecanico_nombre,
                (
                    orden.fecha_ingreso.replace(tzinfo=None)
                    if orden.fecha_ingreso
                    else None
                ),  # Quitar timezone para Excel
                (
                    orden.fecha_entrega_real.replace(tzinfo=None)
                    if orden.fecha_entrega_real
                    else None
                ),
                tiempo_total_taller_horas,
                tiempo_total_pausas_horas,
                tiempo_efectivo_horas,
            ]
        )

        # Formatear fechas
        ws.cell(row=ws.max_row, column=4).number_format = "DD/MM/YYYY HH:MM"
        ws.cell(row=ws.max_row, column=5).number_format = "DD/MM/YYYY HH:MM"

    # 6. Preparar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Reporte_Tiempos_Taller_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"'
    )
    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdministrativo])
def exportar_solicitudes_grua(request):
    """
    Exporta un reporte de todas las solicitudes de grúa
    filtradas por fecha de CREACIÓN de la solicitud.
    """

    # 1. Obtener filtros de fecha
    fecha_inicio_str = request.query_params.get("fecha_inicio", None)
    fecha_fin_str = request.query_params.get("fecha_fin", None)

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Error: Faltan filtros de fecha.", status=400)

    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        fecha_fin_dt = datetime.combine(
            datetime.strptime(fecha_fin_str, "%Y-%m-%d").date(), time.max
        )
    except ValueError:
        return HttpResponse("Error: Formato de fecha inválido.", status=400)

    # 2. Query: Agendamientos que SÍ solicitaron grúa, en el rango de fechas
    solicitudes = (
        Agendamiento.objects.filter(
            solicita_grua=True,
            creado_en__range=[
                fecha_inicio_dt,
                fecha_fin_dt,
            ],  # Filtramos por fecha de solicitud
        )
        .select_related("vehiculo", "chofer_asociado")
        .order_by("-creado_en")
    )

    # 3. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitudes de Grúa"

    # 4. Definir las cabeceras
    headers = [
        "Fecha Solicitud",
        "Patente",
        "Chofer",
        "Dirección de Retiro",
        "Grúa Despachada",
    ]
    ws.append(headers)

    # 5. Añadir los datos
    for item in solicitudes:

        chofer_nombre = (
            item.chofer_asociado.get_full_name() if item.chofer_asociado else "N/A"
        )
        patente_vehiculo = item.vehiculo.patente if item.vehiculo else "N/A"
        estado_grua = "Sí" if item.grua_enviada else "No"

        ws.append(
            [
                item.creado_en.replace(tzinfo=None),  # Quitamos timezone para Excel
                patente_vehiculo,
                chofer_nombre,
                item.direccion_grua,
                estado_grua,
            ]
        )

        # Formatear fecha
        ws.cell(row=ws.max_row, column=1).number_format = "DD/MM/YYYY HH:MM"

    # 6. Preparar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Reporte_Solicitudes_Grua_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"'
    )
    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdministrativo])
def exportar_historial_prestamos(request):
    """
    Exporta un historial de todos los préstamos de llaves (retiros y devoluciones)
    filtrado por la FECHA DE RETIRO.
    """

    # 1. Obtener filtros de fecha
    fecha_inicio_str = request.query_params.get("fecha_inicio", None)
    fecha_fin_str = request.query_params.get("fecha_fin", None)

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Error: Faltan filtros de fecha.", status=400)

    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        fecha_fin_dt = datetime.combine(
            datetime.strptime(fecha_fin_str, "%Y-%m-%d").date(), time.max
        )
    except ValueError:
        return HttpResponse("Error: Formato de fecha inválido.", status=400)

    # 2. Query: Préstamos cuyo RETIRO fue en el rango de fechas
    prestamos = (
        PrestamoLlave.objects.filter(
            fecha_hora_retiro__range=[fecha_inicio_dt, fecha_fin_dt]
        )
        .select_related("llave__vehiculo", "usuario_retira")
        .order_by("-fecha_hora_retiro")
    )

    # 3. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial Préstamos Llaves"

    # 4. Definir las cabeceras
    headers = [
        "Código Llave",
        "Patente",
        "Quién Retiró",
        "Fecha/Hora Retiro",
        "Fecha/Hora Devolución",
    ]
    ws.append(headers)

    # 5. Añadir los datos
    for item in prestamos:

        ws.append(
            [
                item.llave.codigo_interno,
                item.llave.vehiculo.patente,
                item.usuario_retira.get_full_name(),
                item.fecha_hora_retiro.replace(tzinfo=None),
                (
                    item.fecha_hora_devolucion.replace(tzinfo=None)
                    if item.fecha_hora_devolucion
                    else "Aún Prestada"
                ),
            ]
        )

        # Formatear fechas
        ws.cell(row=ws.max_row, column=4).number_format = "DD/MM/YYYY HH:MM"
        if item.fecha_hora_devolucion:
            ws.cell(row=ws.max_row, column=5).number_format = "DD/MM/YYYY HH:MM"

    # 6. Preparar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Reporte_Historial_Llaves_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"'
    )
    wb.save(response)
    return response


# NUEVA VISTA 2: REPORTE INVENTARIO DE LLAVES (PDF)
@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdministrativo])
def exportar_inventario_llaves_pdf(request):
    """
    Genera un reporte PDF (Snapshot) del estado actual de
    todas las llaves en el inventario. No usa filtros de fecha.
    """

    # 1. Query: TODAS las llaves
    llaves = (
        LlaveVehiculo.objects.all()
        .select_related("vehiculo", "poseedor_actual")
        .order_by("vehiculo__patente", "codigo_interno")
    )

    # 2. Preparar el PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()

    # 3. Títulos del PDF
    fecha_actual = timezone.now().strftime("%d/%m/%Y %H:%M:%S")
    elements.append(Paragraph("Reporte de Inventario de Llaves (Pañol)", styles["h1"]))
    elements.append(Paragraph(f"Generado el: {fecha_actual}", styles["Normal"]))
    elements.append(Paragraph(f"Total Llaves: {llaves.count()}", styles["Normal"]))
    elements.append(Paragraph(" ", styles["Normal"]))  # Espacio

    # 4. Preparar datos para la tabla
    data = [["Código Interno", "Patente", "Tipo", "Estado Actual", "Poseedor Actual"]]

    for llave in llaves:

        poseedor = "N/A"
        if llave.estado == LlaveVehiculo.Estado.PRESTADA and llave.poseedor_actual:
            poseedor = llave.poseedor_actual.get_full_name()
        elif llave.estado == LlaveVehiculo.Estado.EN_BODEGA:
            poseedor = "En Pañol"

        data.append(
            [
                llave.codigo_interno,
                llave.vehiculo.patente,
                llave.get_tipo_display(),
                llave.get_estado_display(),
                poseedor,
            ]
        )

    # 5. Crear y Estilizar la Tabla (similar al otro PDF)
    table = Table(data, colWidths=[100, 100, 80, 80, 140])  # Ajustar anchos
    style = TableStyle(
        [
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#2d3748"),
            ),  # Fondo cabecera
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#4a5568")),  # Fondo filas
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]
    )
    table.setStyle(style)
    elements.append(table)

    # 6. Construir el PDF
    doc.build(elements)

    # 7. Crear la respuesta HTTP
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    filename = f"Snapshot_Inventario_Llaves_{timezone.now().strftime('%Y%m%d')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# NUEVA VISTA 1: REPORTE DE FRECUENCIA DE FALLAS (EXCEL)
@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdministrativo])
def exportar_frecuencia_fallas(request):
    """
    Exporta un ranking de los vehículos que más han ingresado al taller,
    filtrado por fecha de INGRESO.
    """

    # 1. Obtener filtros de fecha
    fecha_inicio_str = request.query_params.get("fecha_inicio", None)
    fecha_fin_str = request.query_params.get("fecha_fin", None)

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Error: Faltan filtros de fecha.", status=400)

    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        fecha_fin_dt = datetime.combine(
            datetime.strptime(fecha_fin_str, "%Y-%m-%d").date(), time.max
        )
    except ValueError:
        return HttpResponse("Error: Formato de fecha inválido.", status=400)

    # 2. Query: Agrupar por Vehículo y contar Órdenes (ingresos)
    frecuencia = (
        Orden.objects.filter(fecha_ingreso__range=[fecha_inicio_dt, fecha_fin_dt])
        .values(
            "vehiculo__patente",
            "vehiculo__chofer__first_name",
            "vehiculo__chofer__last_name",
        )
        .annotate(numero_de_ingresos=Count("id"))
        .order_by("-numero_de_ingresos")
    )

    # 3. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Frecuencia de Fallas"

    # 4. Definir las cabeceras
    headers = ["Patente", "Chofer Asignado", "Número de Ingresos al Taller"]
    ws.append(headers)

    # 5. Añadir los datos
    for item in frecuencia:
        chofer_nombre = f"{item['vehiculo__chofer__first_name'] or ''} {item['vehiculo__chofer__last_name'] or ''}".strip()

        ws.append(
            [
                item["vehiculo__patente"],
                chofer_nombre if chofer_nombre else "N/A",
                item["numero_de_ingresos"],
            ]
        )

    # 6. Preparar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Reporte_Frecuencia_Fallas_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"'
    )
    wb.save(response)
    return response


# NUEVA VISTA 2: HOJA DE VIDA DEL VEHÍCULO (PDF)
@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdministrativo])
def exportar_hoja_vida_vehiculo_pdf(request):
    """
    Genera un reporte PDF con el historial completo (Hoja de Vida)
    de un vehículo específico, usando la patente.
    """

    # 1. Obtener filtro de PATENTE
    patente = request.query_params.get("patente", None)
    if not patente:
        return HttpResponse("Error: Debe proporcionar una patente.", status=400)

    # 2. Obtener el vehículo y su historial
    vehiculo = get_object_or_404(Vehiculo, patente=patente)
    ordenes = (
        Orden.objects.filter(vehiculo=vehiculo)
        .prefetch_related("items__producto", "items__servicio")  # Optimización clave
        .order_by("-fecha_ingreso")
    )

    # 3. Preparar el PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        topMargin=72,
        bottomMargin=72,
        leftMargin=72,
        rightMargin=72,
    )
    elements = []

    styles = getSampleStyleSheet()
    styles["h1"].alignment = 1  # Centrado
    styles["h2"].fontSize = 14
    styles["h2"].spaceAfter = 10

    # 4. Títulos del PDF
    fecha_actual = timezone.now().strftime("%d/%m/%Y")
    chofer_nombre = (
        vehiculo.chofer.get_full_name() if vehiculo.chofer else "Sin chofer asignado"
    )

    elements.append(Paragraph("Hoja de Vida del Vehículo", styles["h1"]))
    elements.append(Spacer(1, 24))

    # 5. Datos del Vehículo
    data_vehiculo = [
        [
            "Patente:",
            vehiculo.patente,
            "Marca/Modelo:",
            f"{vehiculo.marca} {vehiculo.modelo}",
        ],
        ["Chofer Actual:", chofer_nombre, "Año:", str(vehiculo.anio)],
    ]
    table_vehiculo = Table(data_vehiculo, colWidths=[100, 150, 100, 150])
    table_vehiculo.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    elements.append(table_vehiculo)
    elements.append(Spacer(1, 24))

    # 6. Iterar sobre CADA ORDEN del vehículo
    for orden in ordenes:
        fecha_ingreso = orden.fecha_ingreso.strftime("%d/%m/%Y %H:%M")
        estado_orden = orden.get_estado_display()

        elements.append(
            Paragraph(
                f"Orden #{orden.id} - Ingreso: {fecha_ingreso} (Estado: {estado_orden})",
                styles["h2"],
            )
        )

        # Falla y Diagnóstico
        falla_cliente = orden.descripcion_falla or "Sin descripción"
        diagnostico_tec = orden.diagnostico_tecnico or "Sin diagnóstico"
        elements.append(
            Paragraph(f"<b>Falla (Cliente):</b> {falla_cliente}", styles["Normal"])
        )
        elements.append(
            Paragraph(
                f"<b>Diagnóstico (Técnico):</b> {diagnostico_tec}", styles["Normal"]
            )
        )
        elements.append(Spacer(1, 12))

        # 7. Tabla de Items (Repuestos y Servicios) para ESTA orden
        items_data = [
            ["Cantidad", "Ítem (Repuesto/Servicio)", "Precio Unit.", "Subtotal"]
        ]

        items_orden = orden.items.all()
        if not items_orden:
            items_data.append(
                ["-", "Esta orden no registró repuestos ni servicios.", "-", "-"]
            )
        else:
            for item in items_orden:
                nombre_item = (
                    item.producto.nombre if item.producto else item.servicio.nombre
                )
                items_data.append(
                    [
                        f"{item.cantidad:.0f}",
                        nombre_item,
                        f"${item.precio_unitario:,.0f}",
                        f"${item.subtotal:,.0f}",
                    ]
                )

        # Estilo de la tabla de ítems
        table_items = Table(items_data, colWidths=[60, 260, 80, 80])
        table_items.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a5568")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("ALIGN", (0, 1), (0, -1), "CENTER"),
                    ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                ]
            )
        )
        elements.append(table_items)
        elements.append(Spacer(1, 24))  # Espacio grande entre órdenes

    # 8. Construir el PDF
    doc.build(elements)

    # 9. Crear la respuesta HTTP
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    filename = f"Hoja_De_Vida_{patente}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
